import typer
import json
import subprocess
import logging
import os
import tarfile
import tempfile
from pathlib import Path
from typing import Optional, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import yaml


def get_logger(name: str = "kdiff") -> logging.Logger:
    """
    Get a logger instance with level set from KDIFF_LOGLEVEL environment variable.

    Args:
        name: The name of the logger

    Returns:
        A configured logger instance
    """
    logger = logging.getLogger(name)

    # Only configure the logger if it hasn't been configured yet
    if not logger.handlers:
        # Get log level from environment variable
        log_level_str = os.getenv("KDIFF_LOGLEVEL", "INFO").upper()

        # Map string levels to logging constants
        level_mapping = {
            "DEBUG": logging.DEBUG,
            "INFO": logging.INFO,
            "WARNING": logging.WARNING,
            "ERROR": logging.ERROR,
            "CRITICAL": logging.CRITICAL,
        }

        log_level = level_mapping.get(log_level_str, logging.INFO)

        # Configure the logger
        logger.setLevel(log_level)

        # Create console handler if none exists
        handler = logging.StreamHandler()
        handler.setLevel(log_level)

        # Create formatter
        formatter = logging.Formatter(
            "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
        )
        handler.setFormatter(formatter)

        # Add handler to logger
        logger.addHandler(handler)

    return logger


def load_checksums(checksums_file: Path) -> Dict[str, str]:
    """
    Load checksums from a checksums.txt
    Expected format: hash filepath (one per line)
    Returns: dict mapping filepath to hash
    """
    checksums = {}
    if checksums_file.exists():
        with open(checksums_file, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):  # Skip empty lines and comments
                    parts = line.split(None, 1)  # Split on whitespace, max 1 split
                    if len(parts) == 2:
                        hash_value, filepath = parts
                        checksums[filepath] = hash_value
    return checksums


def compare_folders(left: Path, right: Path) -> Dict:
    """
    Compare two folders and return the differences as a JSON-serializable dictionary
    """
    # Check if paths are directories
    left_is_dir = left.is_dir()
    right_is_dir = right.is_dir()

    if not (left_is_dir and right_is_dir):
        raise ValueError("Both paths must be directories")

    # Load checksums from both folders
    left_checksums_file = left / "checksums.txt"
    right_checksums_file = right / "checksums.txt"

    left_checksums = load_checksums(left_checksums_file)
    right_checksums = load_checksums(right_checksums_file)

    # Get the list of files from checksums (excluding checksums.txt itself)
    left_files = set(left_checksums.keys())
    right_files = set(right_checksums.keys())

    # Create three lists: left_only, right_only, both
    left_only = list(left_files - right_files)
    right_only = list(right_files - left_files)
    both = list(left_files & right_files)

    # Analyze files in both folders
    identical_files = []
    different_files = []

    for f in sorted(both):
        left_hash = left_checksums.get(f)
        right_hash = right_checksums.get(f)
        if left_hash != right_hash:
            different_files.append(f)
        else:
            identical_files.append(f)

    # Return structured data
    return {
        "left_path": str(left),
        "right_path": str(right),
        "files_only_in_left": sorted(left_only),
        "files_only_in_right": sorted(right_only),
        "identical_files": identical_files,
        "different_files": different_files,
        "summary": {
            "files_only_in_left": len(left_only),
            "files_only_in_right": len(right_only),
            "identical_files": len(identical_files),
            "different_files": len(different_files),
        },
    }


def qsv_diff_different_files(left: Path, right: Path, different_files: list) -> None:
    """
    Process different files and run qsv diff on CSV files
    """
    logger = get_logger()

    # Create diff directory if it doesn't exist
    diff_dir = Path("diff")
    diff_dir.mkdir(exist_ok=True)

    if not different_files:
        logger.info("No different files found.")
        return

    logger.info(f"Found {len(different_files)} different files to process:")

    for file_path in different_files:
        # Check if it's a CSV file
        if not file_path.endswith(".csv"):
            logger.info(f"Skipping non-CSV file: {file_path}")
            continue

        # Construct full paths
        left_file = left / file_path
        right_file = right / file_path

        # Check if both files exist
        if not left_file.exists():
            logger.warning(f"Left file does not exist: {left_file}")
            continue
        if not right_file.exists():
            logger.warning(f"Right file does not exist: {right_file}")
            continue

        # Create output diff file path
        diff_file = diff_dir / f"{Path(file_path).stem}.diff.csv"

        # Run qsv diff command
        cmd = [
            "qsv",
            "diff",
            "--drop-equal-fields",
            "--key",
            "uid",
            # "--sort-columns namespace,name",
            str(left_file),
            str(right_file),
        ]

        logger.info(f"Processing: {file_path}")
        logger.debug(f"Command: {' '.join(cmd)} > {diff_file}")

        try:
            with open(diff_file, "w") as output_file:
                subprocess.run(cmd, stdout=output_file, check=True)
            logger.info(f"✓ Diff saved to: {diff_file}")
        except subprocess.CalledProcessError as e:
            logger.error(f"✗ Error running qsv diff for {file_path}: {e}")
        except Exception as e:
            logger.error(f"✗ Unexpected error processing {file_path}: {e}")

    logger.info(f"Diff processing completed. Results saved in: {diff_dir}")


def create_excel_from_backup(backup_path: Path, output_path: Optional[Path] = None) -> None:
    """
    Extract CSV files from a backup (tar file or folder) and create an Excel file with different sheets
    
    Args:
        backup_path: Path to the backup (tar file or folder)
        output_path: Optional output path for the Excel file
    """
    logger = get_logger()
    
    # Determine output path
    if output_path is None:
        if backup_path.is_file() and backup_path.suffix in ['.tar', '.tar.gz', '.tgz']:
            output_path = backup_path.with_suffix('.xlsx')
        else:
            output_path = backup_path.with_suffix('.xlsx')
    
    logger.info(f"Processing backup: {backup_path}")
    logger.info(f"Output Excel file: {output_path}")
    
    # Check if it's a tar file or folder
    if backup_path.is_file() and backup_path.suffix in ['.tar', '.tar.gz', '.tgz']:
        # Handle tar file
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            
            # Extract the tar file
            logger.info("Extracting backup tar file...")
            with tarfile.open(backup_path, 'r:*') as tar:
                tar.extractall(temp_path)
            
            # Process CSV files from extracted directory
            process_csv_files_to_excel(temp_path, output_path, logger)
    elif backup_path.is_dir():
        # Handle folder directly
        logger.info("Processing backup folder...")
        process_csv_files_to_excel(backup_path, output_path, logger)
    else:
        raise ValueError(f"Invalid backup path: {backup_path}. Must be a tar file (.tar, .tar.gz, .tgz) or a folder.")


def format_json_as_yaml(value):
    """
    Convert JSON string to YAML format for better readability
    
    Args:
        value: String that might contain JSON
    
    Returns:
        Formatted YAML string or original value if not JSON
    """
    if not isinstance(value, str):
        return value
    
    # Try to parse as JSON
    try:
        # Check if it looks like JSON (starts with { or [)
        stripped = value.strip()
        if (stripped.startswith('{') and stripped.endswith('}')) or \
           (stripped.startswith('[') and stripped.endswith(']')):
            parsed = json.loads(value)
            # Convert to YAML with proper formatting and explicit newlines
            yaml_str = yaml.dump(parsed, default_flow_style=False, sort_keys=False, indent=2, allow_unicode=True, width=float("inf"))
            # Ensure we have proper line breaks and trim extra newlines at the end
            return yaml_str.rstrip('\n')
    except (json.JSONDecodeError, yaml.YAMLError):
        pass
    
    return value


def process_csv_files_to_excel(source_path: Path, output_path: Path, logger) -> None:
    """
    Process CSV files from a source path and create an Excel file
    
    Args:
        source_path: Path containing CSV files
        output_path: Output Excel file path
        logger: Logger instance
    """
    # Find all CSV files
    csv_files = list(source_path.rglob("*.csv"))
    logger.info(f"Found {len(csv_files)} CSV files")
    
    if not csv_files:
        raise ValueError("No CSV files found in the backup")
    
    # Create Excel workbook
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Store valid sheets for later sorting
    valid_sheets = []
    
    # Process each CSV file
    for csv_file in csv_files:
        try:
            # Read CSV file
            df = pd.read_csv(csv_file)
            
            # Skip if only has header (no data rows)
            if len(df) == 0:
                logger.info(f"Skipping empty CSV file: {csv_file.name}")
                continue
            
            # Format JSON columns as YAML
            for column in df.columns:
                # Check if column might contain JSON (common column names)
                if any(keyword in column.lower() for keyword in ['json', 'data', 'config', 'spec', 'metadata', 'labels', 'annotations', 'info']):
                    df[column] = df[column].apply(format_json_as_yaml)
            
            # Create sheet name from file name (sanitize for Excel)
            sheet_name = csv_file.stem
            # Excel sheet names have limitations
            sheet_name = sheet_name[:31]  # Max 31 characters
            sheet_name = sheet_name.replace('[', '').replace(']', '').replace('*', '').replace('?', '').replace('/', '').replace('\\', '')
            
            # Store sheet info for later sorting
            valid_sheets.append({
                'name': sheet_name,
                'data': df,
                'row_count': len(df)
            })
            
            logger.info(f"Processed CSV: {csv_file.name} ({len(df)} rows)")
            
        except Exception as e:
            logger.warning(f"Error processing {csv_file.name}: {e}")
            continue
    
    if not valid_sheets:
        raise ValueError("No valid CSV files with data found in the backup")
    
    # Sort sheets alphabetically by name
    valid_sheets.sort(key=lambda x: x['name'].lower())
    
    # Create worksheets in alphabetical order
    for sheet_info in valid_sheets:
        ws = wb.create_sheet(title=sheet_info['name'])
        
        # Add data to worksheet with proper formatting
        for row_idx, row in enumerate(dataframe_to_rows(sheet_info['data'], index=False, header=True), 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Set consistent alignment for all cells (top alignment)
                from openpyxl.styles import Alignment
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Adjust row height for multi-line content
                if isinstance(value, str) and '\n' in value:
                    line_count = value.count('\n') + 1
                    # Set minimum row height based on content
                    min_height = max(15, line_count * 15)  # 15 points per line
                    ws.row_dimensions[row_idx].height = min_height
                    
                    # Force Excel to recognize newlines by setting the cell value explicitly
                    cell.value = str(value)
            
            logger.info(f"Added sheet: {sheet_info['name']} ({sheet_info['row_count']} rows)")
    
    # Save the Excel file
    wb.save(output_path)
    logger.info(f"Excel file saved: {output_path}")


def create_diff_excel(left_backup: Path, right_backup: Path, output_path: Optional[Path] = None) -> None:
    """
    Compare two backups (tar files or folders) and create a colored Excel file showing differences
    
    Args:
        left_backup: Path to the left backup (tar file or folder)
        right_backup: Path to the right backup (tar file or folder)
        output_path: Optional output path for the Excel file
    """
    logger = get_logger()
    
    # Determine output path
    if output_path is None:
        left_name = left_backup.stem
        right_name = right_backup.stem
        output_path = Path(f"diff_{left_name}_vs_{right_name}.xlsx")
    
    logger.info(f"Comparing backups: {left_backup} vs {right_backup}")
    logger.info(f"Output Excel file: {output_path}")
    
    # Helper function to get source path (extract tar if needed)
    def get_source_path(backup_path: Path) -> Path:
        if backup_path.is_file() and backup_path.suffix in ['.tar', '.tar.gz', '.tgz']:
            # Create temporary directory for extraction
            temp_dir = tempfile.mkdtemp()
            temp_path = Path(temp_dir)
            
            logger.info(f"Extracting {backup_path.name}...")
            with tarfile.open(backup_path, 'r:*') as tar:
                tar.extractall(temp_path)
            
            return temp_path
        elif backup_path.is_dir():
            return backup_path
        else:
            raise ValueError(f"Invalid backup path: {backup_path}. Must be a tar file (.tar, .tar.gz, .tgz) or a folder.")
    
    # Get source paths
    left_source = get_source_path(left_backup)
    right_source = get_source_path(right_backup)
    
    try:
        # Compare folders to find different files
        result = compare_folders(left_source, right_source)
        different_files = result.get("different_files", [])
        
        if not different_files:
            logger.info("No different files found between backups")
            return
        
        logger.info(f"Found {len(different_files)} different files to process")
        
        # Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process each different file
        for file_path in different_files:
            if not file_path.endswith('.csv'):
                logger.info(f"Skipping non-CSV file: {file_path}")
                continue
            
            try:
                left_file = left_source / file_path
                right_file = right_source / file_path
                
                if not left_file.exists() or not right_file.exists():
                    logger.warning(f"File not found in one of the backups: {file_path}")
                    continue
                
                # Create sheet name
                sheet_name = Path(file_path).stem
                sheet_name = sheet_name[:31]  # Excel limit
                sheet_name = sheet_name.replace('[', '').replace(']', '').replace('*', '').replace('?', '').replace('/', '').replace('\\', '')
                
                # Run qsv diff
                diff_result = run_qsv_diff(left_file, right_file)
                
                if diff_result is not None:
                    # Create worksheet with colored cells
                    ws = wb.create_sheet(title=sheet_name)
                    create_colored_diff_sheet(ws, diff_result)
                    logger.info(f"Added diff sheet: {sheet_name}")
                
            except Exception as e:
                logger.warning(f"Error processing {file_path}: {e}")
                continue
        
        # Save the Excel file
        wb.save(output_path)
        logger.info(f"Diff Excel file saved: {output_path}")
        
    finally:
        # Clean up temporary directories if they were created
        if left_source != left_backup:
            import shutil
            shutil.rmtree(left_source, ignore_errors=True)
        if right_source != right_backup:
            import shutil
            shutil.rmtree(right_source, ignore_errors=True)


def run_qsv_diff(left_file: Path, right_file: Path) -> Optional[pd.DataFrame]:
    """
    Run qsv diff command and return the result as a DataFrame
    
    Args:
        left_file: Path to left CSV file
        right_file: Path to right CSV file
    
    Returns:
        DataFrame with diff results or None if error
    """
    logger = get_logger()
    
    cmd = [
        "qsv",
        "diff",
        "--drop-equal-fields",
        str(left_file),
        str(right_file),
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        if result.stdout.strip():
            # Parse CSV output
            from io import StringIO
            df = pd.read_csv(StringIO(result.stdout))
            return df
        else:
            logger.info("No differences found")
            return None
    except subprocess.CalledProcessError as e:
        logger.error(f"Error running qsv diff: {e}")
        return None


def create_colored_diff_sheet(ws, df: pd.DataFrame) -> None:
    """
    Create a worksheet with colored cells based on diff results
    
    Args:
        ws: Worksheet to populate
        df: DataFrame with diff results
    """
    from openpyxl.styles import PatternFill
    
    # Define colors
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red for deletions
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green for additions
    orange_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")  # Light orange for modifications
    
    # Add headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = cell.font.copy(bold=True)
    
    # Add data with colors
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Color based on diff type (first column usually indicates + or -)
            if col_idx == 1:
                if str(value).startswith('-'):
                    cell.fill = red_fill
                elif str(value).startswith('+'):
                    cell.fill = green_fill
                else:
                    cell.fill = orange_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
